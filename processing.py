"""Convert the frozen product catalog into an Excel-friendly attribute table.

Usage:
    python processing.py
    python processing.py --output data/catalog_attributes.xlsx --limit 1000

The workbook has one product per row.  The ten conversational attributes are
derived from catalog metadata, so a blank value means no confident match was
found; it does not mean the product lacks that property.
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MATERIALS = (
    "cotton", "polyester", "nylon", "leather", "wool", "spandex", "silk",
    "rayon", "linen", "denim", "suede", "cashmere", "rubber", "fabric",
)
COLORS = (
    "black", "white", "blue", "navy", "cyan", "teal", "red", "pink", "green",
    "brown", "gray", "grey", "purple", "yellow", "orange", "beige", "tan",
    "gold", "silver", "ivory", "cream", "khaki", "multicolor",
)
STYLE_WORDS = (
    "casual", "formal", "classic", "vintage", "sport", "athletic", "boho",
    "fashion", "slim fit", "regular fit", "relaxed fit", "oversized", "elegant",
)
USE_CASE_WORDS = (
    "hiking", "running", "walking", "gym", "workout", "yoga", "swimming",
    "winter", "outdoor", "work", "school", "travel", "wedding", "party",
    "beach", "rain", "ski", "cycling", "sleep",
)
SIZE_RE = re.compile(
    r"\b(?:size|sizes|small|medium|large|x-small|x-large|xxl|xxxl|\d{1,2}(?:\.\d)?\s*(?:us|uk|eu))\b",
    re.IGNORECASE,
)


def flatten(value: object) -> str:
    """Turn nested catalog fields into readable, single-cell text."""
    if value is None:
        return ""
    if isinstance(value, dict):
        return "; ".join(f"{key}: {item}" for key, item in value.items() if item not in (None, "", []))
    if isinstance(value, list):
        return "; ".join(str(item) for item in value if item not in (None, ""))
    return str(value)


def matched_terms(text: str, terms: tuple[str, ...]) -> str:
    lowered = text.lower()
    return "; ".join(term for term in terms if re.search(rf"\b{re.escape(term)}\b", lowered))


def extract_size(text: str) -> str:
    found = list(dict.fromkeys(match.group(0) for match in SIZE_RE.finditer(text)))
    return "; ".join(found)


def clip(value: str, limit: int = 32000) -> str:
    """Keep data inside Excel's 32,767-character cell limit."""
    return value if len(value) <= limit else value[: limit - 3] + "..."


def product_row(product: dict) -> list[object]:
    title = str(product.get("title") or "")
    categories = flatten(product.get("categories"))
    features = flatten(product.get("features"))
    details = flatten(product.get("details"))
    description = flatten(product.get("description"))
    searchable = " ".join((title, categories, features, details, description, str(product.get("store") or "")))
    return [
        str(product.get("parent_asin") or ""),
        clip(title),
        clip(categories),                         # category
        matched_terms(searchable, MATERIALS),     # material
        matched_terms(searchable, COLORS),        # color
        extract_size(searchable),                 # size
        matched_terms(searchable, STYLE_WORDS),   # style
        str(product.get("store") or ""),        # brand (catalog's closest field)
        product.get("price"),                    # budget
        clip(features),                           # feature
        matched_terms(searchable, USE_CASE_WORDS),# use_case
        clip(details),                            # other
    ]


def write_workbook(catalog_path: Path, output_path: Path, limit: int | None) -> int:
    headers = [
        "parent_asin", "title", "category", "material", "color", "size", "style",
        "brand", "budget", "feature", "use_case", "other",
    ]
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Catalog attributes")
    sheet.append(headers)
    count = 0
    with catalog_path.open(encoding="utf-8") as source:
        for line in source:
            if not line.strip():
                continue
            sheet.append(product_row(json.loads(line)))
            count += 1
            if limit is not None and count >= limit:
                break
    workbook.save(output_path)
    return count


def main() -> None:
    print("Exporting catalog metadata into ten conversational attributes...")
    parser = argparse.ArgumentParser(description="Export catalog metadata into ten conversational attributes.")
    parser.add_argument("--catalog", type=Path, default=Path("data/catalog.jsonl"))
    parser.add_argument("--output", type=Path, default=Path("data/catalog_attributes.xlsx"))
    parser.add_argument("--limit", type=int, default=None, help="Optional number of products to export.")
    args = parser.parse_args()
    if args.limit is not None and args.limit < 1:
        parser.error("--limit must be positive")
    args.output.parent.mkdir(parents=True, exist_ok=True)
    rows = write_workbook(args.catalog, args.output, args.limit)
    print(f"Wrote {rows:,} products to {args.output}")


if __name__ == "__main__":
    main()
